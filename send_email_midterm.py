# -*- coding: UTF-8 -*-
'''
A demo code to deliver the grades to students in NSD.
author: Minghan Yang
date:   13 May, 2019.
'''
import smtplib,ssl
from email.mime.text import MIMEText
from openpyxl import load_workbook
import argparse
import os

parser = argparse.ArgumentParser(description='发送考试成绩至同学邮箱')
parser.add_argument('--stat_dir',default='test.xlsx',help='Please input a xlsx file!')
parser.add_argument('--sender_email',default='yangorwell@163.com',help='Please input your email adress!')
parser.add_argument('--send_server',default='smtp.163.com',help='Please input your email server, default 163 smtp server.')
parser.add_argument('--save_log_dir',default='./log/',help='Save the log in this dir.')
args = parser.parse_args()
assert os.path.splitext(args.stat_dir)[-1] == '.xlsx' 

print("Loading data from %s" %(args.stat_dir))
info = load_workbook(args.stat_dir)
table = info[info.sheetnames[0]]
port = 465

smtp_server = args.send_server
sender_email = args.sender_email

password = input('please input your password')

successful_receiver = []
unsuccessful_receiver = []

num  = 0
context = ssl.create_default_context()
for [a,b,c] in table.iter_rows(min_row=2):
    print()
    # server.login(sender_email, password)
    
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        content = """{name}同学，您好。
(请勿回复本邮箱，邮件为自动发送)
    """
        server.connect(smtp_server)
        server.login(sender_email, password)
        num = num + 1
        print("Sending email to %d th student, his/her email is %s ....." %(num,c.value))
        # print(content.format(name=a.value,grade=b.value))
        message = MIMEText(content.format(name=a.value,grade=b.value),'plain','utf-8')
        message['Subject'] = '国发院概率统计期中成绩分数{}'.format(b.value)
        message['From'] = sender_email+'<'+sender_email+'>'
        message['To'] = c.value+'<'+c.value+'>'

        
        try:
            server.sendmail(sender_email, c.value, message.as_string())

            print('mail has been send successfully.')
            successful_receiver.append([a.value,c.value])
        except smtplib.SMTPException as e:
            print(e)
            print('unsuccess sending')
            unsuccessful_receiver.append([a.value,c.value])
        # print()

print(successful_receiver,unsuccessful_receiver)


if not os.path.exists(args.save_log_dir):
    os.makedirs(args.save_log_dir)


with open(args.save_log_dir+'successful_receiver.log','w') as f:
    for line in successful_receiver:
        # print(line)
        f.writelines(line)
        f.writelines('\n')

print("Successful email-receiever log has been saved in %s" %(args.save_log_dir))

with open(args.save_log_dir+'unsuccessful_receiver.log','w') as f:
    for line in unsuccessful_receiver:
        # print(line)
        f.writelines(line)
        f.writelines('\n')
print("Unsuccessful email-receiever log has been saved in %s, please contact to these students!" %(args.save_log_dir))
